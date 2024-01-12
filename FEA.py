# -*- coding: utf-8 -*-
"""
Created on Mon Jan  8 11:21:56 2024

@author: baret
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Jul 17 17:00:45 2023

@author: baret
"""

# -*- coding: utf-8 -*-
"""
Created on Wed Jun 21 11:17:03 2023

@author: baret
"""

# -*- coding: utf-8 -*-
"""
Created on Tue Jun 20 12:24:24 2023

@author: baret
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Jan 23 14:12:05 2023

@author: baret
"""
import pandas as pd

import numpy as np
import time
import os
import itertools
import xlsxwriter
from openpyxl import load_workbook



def calc_matrice_transition(Tri,NbF,Proba, Niveaux,mu ):
    #(Niveaux)
   # print(Proba)
    Q=np.zeros(((NbF+2),(NbF+2)))
    j1=(int(Tri[0]))
    Q[0,0]=mu
    Q[0,j1]= 1-mu

    for i in range (NbF):

        jinit=int(Tri[i])

        #print(Proba)
        #print(Proba)
        if i < NbF-1:
            j=int(Tri[i+1])
            Q[jinit, 0]=Proba[jinit-1,Niveaux[jinit-1]]
            Q[jinit,j]=1-Proba[jinit-1,Niveaux[jinit-1]]
            
        else :
           
            # print(j)
            # print(j3)
             Q[jinit, 0]=Proba[jinit-1,Niveaux[jinit-1]]
             Q[jinit,NbF+1]=1-Proba[jinit-1,Niveaux[jinit-1]]
            
  #   print(Q)
    P=0
    P2=np.zeros(NbF)
    for k in range( NbF) :
        P=P+(1-Proba[k,Niveaux[k]])
        P2[k]=P2[k]+(Proba[k,Niveaux[k]])
        for kk in range(NbF):
            if kk != k : 
                P2[k]=P2[k]+(1-Proba[kk,Niveaux[kk]])
    for l in range(0,NbF):
        Q[NbF+1,l+1]=(P2[l])/(NbF+P)
    return Q
    
def calc_matrice_transition2(Tri,NbF,Proba, Niveaux,mu):
        Q=np.zeros(((NbF*2)+1,(NbF*2)+1))
        Q[0,0]=mu
        
        #print(self.Tri)
        j3=int((Tri[0]*2))
        
        Q[0,j3]=1-mu
        for i in range (NbF):
            j=int(((Tri[i])*2)-1)
            jinit=int(Tri[i])
           # print('test')
            #print(jinit-1)
            #print(Proba)
            #print(Niveaux)
            #print(Proba)
            if i < NbF-1:
                
                #print(Proba)
                Q[j, 0]=Proba[jinit-1,int(Niveaux[jinit-1])]
                Q[j,j+1]=1-Proba[jinit-1,int(Niveaux[jinit-1])]
                Q[j+1,j]=Proba[jinit-1,int(Niveaux[jinit-1])]
                j2=int((Tri[i+1]*2))
                Q[j+1,j2]=1-Proba[jinit-1,int(Niveaux[jinit-1])]
            else :
               
               # print(j)
               # print(j3)
                Q[j, 0]=Proba[jinit-1,int(Niveaux[jinit-1])]
                Q[j,j+1]=1-Proba[jinit-1,int(Niveaux[jinit-1])]
                Q[j+1,j]=Proba[jinit-1,int(Niveaux[jinit-1])]
                Q[j+1, j3]=1-Proba[jinit-1,int(Niveaux[jinit-1])]
        return Q
    
def calc_matrice_transition0(Tri,NbF,proba, Niveaux,mu):
        Q=np.zeros((NbF)+1)
        
        l=1
        for k in range (NbF): 
            t1=1
            for tau in range(k-1):
                t1=t1*(1-proba[k,Niveaux[k]])
            t=t1*(proba[k,Niveaux[k]])
            l=l*(1-proba[k,Niveaux[k]])
            Q[k]=t
        Q[NbF]=l
        return Q

def calc_steady_state2(p):
    dim = p.shape[0]
    q = (p-np.eye(dim))
    ones = np.ones(dim)
    q = np.c_[q,ones]
    QTQ = np.dot(q, q.T)
    bQT = np.ones(dim)
    return(np.linalg.solve(QTQ,bQT))



def calc_sol_10(nbL, B_lim):

    L=[i for i in range(nbL)]


    N1=(list(itertools.combinations(L, 1)))
    N2=list(itertools.combinations(L, 1)) 
    N3=(list(itertools.combinations(L, 1)))
    N4=(list(itertools.combinations(L, 1)))
    N5=(list(itertools.combinations(L, 1)))
    N6=(list(itertools.combinations(L, 1)))
    N7=(list(itertools.combinations(L, 1)))
    N8=(list(itertools.combinations(L, 1)))
    N9=(list(itertools.combinations(L, 1)))
    N10=(list(itertools.combinations(L, 1)))
    Sol=[]
    
    for k1 in range(0,nbL):
        for k2 in range(0,nbL):
            for k3 in range(0,nbL):
                for k4 in range(0,nbL):
                    for k5 in range(0,nbL):
                        for k6 in range(0,nbL):
                            for k7 in range(0,nbL):
                                for k8 in range(0,nbL):
                                    for k9 in range(0,nbL):
                                        for k10 in range(0,nbL):
                                                New=N1[k1]+N2[k2]+N3[k3]+N4[k4]+N5[k5]+N6[k6]+N7[k7]+N8[k8]+N9[k9]+ N10[k10]
                                                B=calc_budget(New)
                                                if B <= B_lim:
                                                    Sol.append(New)
    return Sol 




def calc_nb_sol(sol, B): 
    nb=0
    for i in range(len(sol)):
        b=calc_budget(sol[i])
        if b<=B:
            nb=nb+1
    return nb
    
def calc_ND(sol, B_test, budget):
    iB=[]
    xB = []
    yB = []
    ND=[]
    for k in range(len(sol)):
        #print(budget[k])
        if budget[k] <= B_test:
            iB.append(k)
            xB.append(sol[k][1])
           
    m=np.min(xB)
    mI=np.argmin(xB)
    ND=[iB[mI],m]
    print(ND)
    return ND

def calc_sol_B(sol):

    ND=[]
  #  print(sol)

    for k10 in range(len(sol)):
        test = True
        for k11 in range(len(sol)):
            if sol[k11][1] < sol[k10][1] and sol[k11][2] < sol[k10][2]:
                    test = False

        if test == True:
            ND.append([sol[k10][0],sol[k10][1],sol[k10][2]])
  #  print(xB)
    return ND


def analyse_res(fichierAnalyse, SS,Dem,NbF,NbC):
    Tensions=np.zeros(NbF)
    
    for k in range(NbF):
        for i in range(NbC):
            Tensions[k]=SS[i][k+1]*Dem[i]+Tensions[k]
            
    workbook = xlsxwriter.Workbook(fichierAnalyse)
    worksheet = workbook.add_worksheet('Result')
    for k in range(NbF):
            worksheet.write_number(k+1,2,k)
            worksheet.write_number(k+1,3,Tensions[k])
    workbook.close()
            
def write_sol(fichierSol,sol,Niveaux):
    workbook = xlsxwriter.Workbook(fichierSol)
    worksheet = workbook.add_worksheet('Result')
    
    for k1 in range(len(sol)):
        B=0
        # for k2 in range(len(Niveaux[k1])):
        #     B=B+10*Niveaux[k1][k2]
        worksheet.write_number(k1+1,2,sol[k1][0])
        worksheet.write_number(k1+1,3,sol[k1][1])
        worksheet.write_number(k1+1,4,sol[k1][2])
        worksheet.write_number(k1+1,5,sol[k1][3])
    workbook.close()
    
def write_sol2(fichierSol,sol,Niveaux):
    workbook = xlsxwriter.Workbook(fichierSol)
    worksheet = workbook.add_worksheet('Result')
    
    B=0
        # for k2 in range(len(Niveaux[k1])):
        #     B=B+10*Niveaux[k1][k2]
        
    for i in range(len(sol)): 
        worksheet.write_number(i+2,2,sol[i][0])
        worksheet.write_number(i+2,3,sol[i][1])
        worksheet.write_number(i+2,4,sol[i][2])
        for k in range(len(Niveaux[sol[i][0]])):

            worksheet.write_number(i+2,k+6,Niveaux[sol[i][0]][k])

    workbook.close()

def write_comb(fichierComb,comb):
    workbook = xlsxwriter.Workbook(fichierComb)
    worksheet = workbook.add_worksheet('Result')
    
    for k1 in range(len(comb)):
        for k2 in range(len(comb[1])):
            worksheet.write_number(k1+1,k2+1,comb[k1][k2])
    workbook.close()
    
def write_ND2(filesol,Sol):
    workbook = load_workbook(filesol)
    worksheetP=workbook['Sheet1']
    c=worksheetP.cell(row=2,column=2)
    c.value='Isntance'
    c2=worksheetP.cell(row=2,column=3)
    c2.value='O1'
    c3=worksheetP.cell(row=2,column=4)
    c3.value='O2'
    c4=worksheetP.cell(row=2,column=5)
    c4.value='B'
    ligne=0
    for i in range(len(Sol)) :
            ligne=ligne+1
            #print(i)
            #print(Sol[i])
            for j in range(len(Sol[i])):
                ligne=ligne+1
                c9=worksheetP.cell(row=ligne,column=1)
                c9.value=Sol[i][j][0]
                c5=worksheetP.cell(row=ligne,column=2)
                c5.value=i
                c6=worksheetP.cell(row=ligne,column=3)
                c6.value=Sol[i][j][1]
                c7=worksheetP.cell(row=ligne,column=4)
                c7.value=Sol[i][j][2]
                c8=worksheetP.cell(row=ligne,column=5)
               # print(budget[i])
                c8.value=Sol[i][j][3]
    workbook.save(filesol)
    return 

def calc_budget(w):
    B=0
    for i in range(len(w)):
        B=B+10*w[i]
    return B

def main(fichier,NbC,NbF,NbL, limite_B, Budg): 
 
    NbSOl=(NbL**NbF)
    NbS=int(NbSOl)
    print(NbS)

    T=pd.read_excel(fichier,sheet_name="Tri", index_col=0 )
    Tri =T.to_numpy()
    P=pd.read_excel(fichier, sheet_name="Proba", index_col=0)
    ProbaF=P.to_numpy()
   
    Dem=pd.read_excel(fichier, sheet_name="Demandes", index_col=0)
    Demande=Dem.to_numpy()

    #print(Demande)
    Dist=pd.read_excel(fichier,sheet_name="Distances", index_col=0 )
    D =Dist.to_numpy()

    #!!!!!!!!!!!!!!!!!!!!!!
    Niveaux=calc_sol_10(NbL,limite_B*Budg)
    #!!!!!!!!!!!!!!!!!!!!!!!!!!
  #  print(NbS)
    #Sol_obj=np.zeros((NbS,2))
    Sol_obj=[]
    #print('sol ob')
    #print(len(Sol_obj))
    Budget=[]
    nb_reel=0
    print(len(Niveaux))
   # print(ProbaF)

    for w in range(len(Niveaux)):
   # for w in range(1):
        print(w)
        O1=0
        O2=0
        Steady_state=[]
        B=calc_budget(Niveaux[w])
      #  print(B)
      #  print(limite_B*Budg)
        Budget.append(B)
        if B <=limite_B*Budg:
            Obj1=0
            Obj2=0
            nb_reel=nb_reel+1
           # print('Solution ' + str(w))
           
            for i in range(NbC): 
                
                SS=calc_matrice_transition0(Tri[i], NbF, ProbaF, Niveaux[w], 0.8)
                
                #print(Q)
                #SS=calc_steady_state2(Q)
                
                Steady_state.append(SS)
          
           
            somme=0

           # print(len(D))
           # print(Steady_state)
            for i2 in range(NbC):
                Proba1=0
                Proba2=0
                for k in range(NbF):
                    somme=somme+1
                    
                    
                   # print((Steady_state[i2][k2]*D[i2,k])*Demande[i2,0])
                  #  print('test')
                   # print(Steady_state[i2][k2]*D[i2,k]*(Demande[i2]))
                    Proba1=Proba1+(Steady_state[i2][k]*D[i2][k])
               
                    
                    # if Proba2 >= 1: 
                    #     print("Probleme !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
                    #     print((Steady_state[i2]))
                    #     print((Steady_state[i2][k3]))
                    #     print(i2)    
                    #     print(k3)
                    #     print(Proba2)
                k3=NbF

                Proba2=(Steady_state[i2][k3])
                Obj1=Proba1*Demande[i2][0]

                Obj2=Obj2+Proba2*Demande[i2][0]
                # print("Proba1")
                # print(Obj1)
                # print("proba2")
                # print(Obj2) 
          #  print(O1)
            O1=Obj1
            O2=Obj2
            #print(somme)
            # Sol_obj[w,0]=O1
            # Sol_obj[w,1]=O2
        #print((O1/Dmax))
        Sol_obj.append([w,O1,O2,B])

    return Sol_obj,Niveaux, NbS, D, Steady_state,Budget, nb_reel


NbC=405
NbF=10
NbL=3
limite_B=0.5
B=(NbL-1)*10*NbF
print(B)

st = time.process_time()
fichier=f'Aube/Aube_fin.xlsx'

sol, Niveaux, NbS, D, Steady,budget, nb_reel =main(fichier,NbC,NbF,NbL,limite_B,B)
#print(budget)
        # print(sol)
#         #print(sol.size)
print(sol)
et = time.process_time()
fichierSol=f'Aube_{NbC}_{NbF}_{NbL}-2_non_recurrent.xlsx'
ND=calc_sol_B(sol)
print(ND)
write_sol2(fichierSol,ND,Niveaux) 

etND=time.process_time()
res = et - st


    # get execution time


print('CPU Execution time:', res, 'seconds')  

 
